"""
📘 보험 약관 PDF 전처리기
작업계획서 Step 1: PDF에서 텍스트+표 추출하여 페이지별로 저장
"""

import fitz  # PyMuPDF
import pdfplumber
import json
import os
from typing import List, Dict, Any
import pandas as pd
import camelot
import re

class PDFPreprocessor:
    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.filename = os.path.splitext(os.path.basename(pdf_path))[0]
    
    def clean_text(self, text):
        """불필요한 공백 및 개행문자 삭제"""
        if not text:
            return ""
        # 연속된 공백을 하나로
        text = re.sub(r' +', ' ', text)
        # 연속된 개행을 하나로
        text = re.sub(r'\n\s*\n+', '\n\n', text)
        # 앞뒤 공백 제거
        text = text.strip()
        return text

    def clean_cell_value(self, value):
        """셀 값에서 Excel에서 허용하지 않는 문자 제거"""
        if value is None or pd.isna(value):
            return ""
        
        str_value = str(value)
        
        # 제어 문자 제거
        cleaned = ""
        for char in str_value:
            code = ord(char)
            if code == 0x09 or code == 0x0A or code == 0x0D:
                cleaned += char
            elif 0x20 <= code <= 0x7E or code >= 0xA0:
                cleaned += char
        
        return cleaned.strip()

    def fill_merged_cells(self, df):
        """DataFrame에서 병합된 셀 채우기"""
        # 복사본 생성
        filled_df = df.copy()
        
        # 1. 가로 방향 병합 처리 (왼쪽 -> 오른쪽)
        for col in filled_df.columns:
            filled_df[col] = filled_df[col].ffill()
        
        # 2. 세로 방향 병합 처리 (위 -> 아래)
        filled_df = filled_df.ffill()
        
        # 3. 남은 NaN을 빈 문자열로
        filled_df = filled_df.fillna("")
        
        # 4. 빈 문자열 채우기
        for col in filled_df.columns:
            for idx in filled_df.index:
                if filled_df.at[idx, col] == "":
                    # 왼쪽 셀 확인
                    col_idx = filled_df.columns.get_loc(col)
                    if col_idx > 0:
                        prev_col = filled_df.columns[col_idx - 1]
                        if filled_df.at[idx, prev_col] != "":
                            filled_df.at[idx, col] = filled_df.at[idx, prev_col]
                    # 위 셀 확인
                    if filled_df.at[idx, col] == "" and idx > 0:
                        if filled_df.at[idx - 1, col] != "":
                            filled_df.at[idx, col] = filled_df.at[idx - 1, col]
        
        return filled_df

    def extract_tables_with_camelot(self, page_num):
        """Camelot을 사용하여 특정 페이지에서 표 추출"""
        tables_text = ""
        tables_count = 0
        
        try:
            # 방법 1: lattice mode (테두리가 있는 표)
            try:
                tables_lattice = camelot.read_pdf(
                    self.pdf_path,
                    pages=str(page_num),
                    flavor='lattice',
                    line_scale=40,
                    copy_text=['v', 'h']
                )
                print(f"     Lattice mode: {len(tables_lattice)}개 표 발견")
            except Exception as e:
                print(f"     Lattice mode 오류: {e}")
                tables_lattice = []
            
            # 방법 2: stream mode (테두리가 없는 표)
            try:
                tables_stream = camelot.read_pdf(
                    self.pdf_path,
                    pages=str(page_num),
                    flavor='stream',
                    edge_tol=50,
                    row_tol=10,
                    column_tol=10
                )
                print(f"     Stream mode: {len(tables_stream)}개 표 발견")
            except Exception as e:
                print(f"     Stream mode 오류: {e}")
                tables_stream = []
            
            # 두 방법 중 더 많은 표를 찾은 것 선택
            page_tables = tables_lattice if len(tables_lattice) >= len(tables_stream) else tables_stream
            mode_used = "lattice" if len(tables_lattice) >= len(tables_stream) else "stream"
            
            if len(page_tables) > 0:
                print(f"     선택된 모드: {mode_used} ({len(page_tables)}개 표)")
                
                # 표를 x 좌표(왼쪽->오른쪽)로 정렬
                tables_with_position = []
                for table in page_tables:
                    df = table.df
                    
                    # 빈 표 건너뛰기
                    if df.empty or df.shape[0] == 0:
                        continue
                    
                    # 표의 x 좌표 (왼쪽 위치)
                    x_position = table._bbox[0] if hasattr(table, '_bbox') else 0
                    
                    tables_with_position.append({
                        'table': table,
                        'x_position': x_position,
                        'df': df
                    })
                
                # x 좌표로 정렬 (왼쪽부터)
                tables_with_position.sort(key=lambda t: t['x_position'])
                
                # 정렬된 순서로 텍스트 변환
                for i, item in enumerate(tables_with_position):
                    df = item['df']
                    
                    # 병합된 셀 처리
                    filled_df = self.fill_merged_cells(df)
                    
                    tables_text += f"\n\n[표 {i+1}]\n"
                    
                    # DataFrame을 텍스트로 변환
                    for row_idx in range(len(filled_df)):
                        row_data = []
                        for col_idx in range(len(filled_df.columns)):
                            cell_value = filled_df.iloc[row_idx, col_idx]
                            cleaned_value = self.clean_cell_value(cell_value)
                            if cleaned_value.strip():
                                row_data.append(cleaned_value.strip())
                        
                        if row_data:
                            tables_text += " | ".join(row_data) + "\n"
                    
                    tables_count += 1
                    
        except Exception as e:
            print(f"   Camelot 표 추출 오류 (페이지 {page_num}): {e}")
        
        return tables_text, tables_count
        
    def extract_text_and_tables(self) -> List[Dict[str, Any]]:
        """PDF에서 텍스트와 표를 추출하여 페이지별로 반환"""
        print(f"PDF 처리 시작: {self.pdf_path}")
        
        # PyMuPDF로 텍스트 추출
        doc = fitz.open(self.pdf_path)
        pages_data = []
        
        for page_num in range(len(doc)):
            print(f"   페이지 {page_num + 1} 처리 중...")
            
            # PyMuPDF로 텍스트 추출
            page = doc[page_num]
            text = page.get_text()
            
            # Camelot으로 표 추출
            tables_text, tables_count = self.extract_tables_with_camelot(page_num + 1)
            if tables_count > 0:
                print(f"     총 {tables_count}개 표를 텍스트로 변환 완료")
                print(f"     표 텍스트 길이: {len(tables_text)}자")
            else:
                print(f"     표 없음")
            
            # 텍스트와 표 결합
            combined_text = text.strip()
            if tables_text.strip():
                combined_text += "\n" + tables_text.strip()
            
            pages_data.append({
                "page": page_num + 1,
                "text": combined_text,
                "source": os.path.basename(self.pdf_path)
            })
        
        doc.close()
        print(f"총 {len(pages_data)}페이지 처리 완료")
        return pages_data
    
    def save_to_files(self, pages_data: List[Dict[str, Any]]):
        """추출된 데이터를 파일로 저장"""
        # 출력 폴더 준비
        os.makedirs("processed_data", exist_ok=True)

        # 1. JSON 파일 저장 (벡터 DB용)
        json_path = os.path.join("processed_data", f"{self.filename}_pages.json")
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(pages_data, f, ensure_ascii=False, indent=2)
        print(f"JSON 저장: {json_path}")
        
        # 2. TXT 파일 저장 (검증용)
        txt_path = os.path.join("processed_data", f"{self.filename}_extracted.txt")
        with open(txt_path, 'w', encoding='utf-8') as f:
            for page_data in pages_data:
                f.write(f"=== 페이지 {page_data['page']} ===\n")
                f.write(page_data['text'])
                f.write("\n\n")
        print(f"TXT 저장: {txt_path}")

        # 3. 표를 엑셀로 저장 (Camelot 사용)
        try:
            import time
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            timestamp = int(time.time())
            xlsx_path = os.path.join("processed_data", f"{self.filename}_tables_{timestamp}.xlsx")
            
            wb = Workbook()
            wb.remove(wb.active)
            
            # 스타일 정의
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border_side = Side(style='thin', color='000000')
            border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
            
            sheet_counter = 1
            for page_data in pages_data:
                page_num = page_data['page']
                
                # Camelot으로 표 추출
                tables_text, tables_count = self.extract_tables_with_camelot(page_num)
                
                if tables_count > 0:
                    # Camelot으로 실제 DataFrame 추출
                    try:
                        # lattice mode 시도
                        tables_lattice = camelot.read_pdf(
                            self.pdf_path,
                            pages=str(page_num),
                            flavor='lattice',
                            line_scale=40,
                            copy_text=['v', 'h']
                        )
                        
                        # stream mode 시도
                        tables_stream = camelot.read_pdf(
                            self.pdf_path,
                            pages=str(page_num),
                            flavor='stream',
                            edge_tol=50,
                            row_tol=10,
                            column_tol=10
                        )
                        
                        # 더 많은 표를 찾은 방법 선택
                        page_tables = tables_lattice if len(tables_lattice) >= len(tables_stream) else tables_stream
                        
                        # 표를 x 좌표로 정렬
                        tables_with_position = []
                        for table in page_tables:
                            df = table.df
                            if df.empty or df.shape[0] == 0:
                                continue
                            x_position = table._bbox[0] if hasattr(table, '_bbox') else 0
                            tables_with_position.append({'table': table, 'x_position': x_position, 'df': df})
                        
                        tables_with_position.sort(key=lambda t: t['x_position'])
                        
                        # 각 표를 엑셀 시트로 저장
                        for t_idx, item in enumerate(tables_with_position):
                            df = item['df']
                            
                            # 병합된 셀 처리
                            filled_df = self.fill_merged_cells(df)
                            
                            # 시트 생성
                            sheet_name = f"P{page_num}_T{t_idx+1}"[:31]
                            ws = wb.create_sheet(title=sheet_name)
                            
                            # 데이터 작성
                            for i in range(len(filled_df)):
                                for j in range(len(filled_df.columns)):
                                    cell_value = filled_df.iloc[i, j]
                                    cleaned_value = self.clean_cell_value(cell_value)
                                    
                                    cell = ws.cell(row=i + 1, column=j + 1, value=cleaned_value)
                                    
                                    # 첫 행은 헤더
                                    if i == 0:
                                        cell.fill = header_fill
                                        cell.font = header_font
                                    else:
                                        cell.font = Font(size=10)
                                    
                                    cell.border = border
                                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                            
                            # 열 너비 조정
                            for col in range(1, len(filled_df.columns) + 1):
                                ws.column_dimensions[get_column_letter(col)].width = 25
                            
                            sheet_counter += 1
                            
                    except Exception as e:
                        print(f"   페이지 {page_num} 엑셀 저장 오류: {e}")
                        continue
            
            wb.save(xlsx_path)
            print(f"엑셀 저장: {xlsx_path}")
            
        except Exception as e:
            print(f"엑셀 저장 중 오류: {e}")
        
        # 3. 통계 출력
        total_chars = sum(len(page['text']) for page in pages_data)
        print(f"통계:")
        print(f"   - 총 페이지: {len(pages_data)}")
        print(f"   - 총 문자수: {total_chars:,}")
        print(f"   - 평균 페이지당: {total_chars // len(pages_data):,}자")

def process_all_pdfs_in_source():
    """source 폴더의 모든 PDF 파일을 처리"""
    import glob
    
    print("source 폴더의 모든 PDF 파일 처리")
    print("=" * 60)
    
    # source 폴더의 모든 PDF 파일 찾기
    pdf_files = glob.glob("source/*.pdf")
    
    if not pdf_files:
        print("source 폴더에 PDF 파일이 없습니다.")
        return
    
    print(f"발견된 PDF 파일: {len(pdf_files)}개")
    for pdf_file in pdf_files:
        print(f"   - {pdf_file}")
    
    # 모든 PDF 데이터를 저장할 리스트
    all_pages_data = []
    
    # 각 PDF 파일 처리
    for pdf_file in pdf_files:
        print(f"\n처리 중: {pdf_file}")
        
        try:
            # PDF 전처리기로 데이터 추출
            processor = PDFPreprocessor(pdf_file)
            pages_data = processor.extract_text_and_tables()
            
            if pages_data:
                print(f"   {len(pages_data)}페이지 추출 완료")
                all_pages_data.extend(pages_data)
                
                # 개별 파일 저장
                processor.save_to_files(pages_data)
            else:
                print(f"   데이터 추출 실패")
                
        except Exception as e:
            print(f"   처리 오류: {str(e)}")
            continue
    
    if not all_pages_data:
        print("처리된 데이터가 없습니다.")
        return
    
    print(f"\n총 처리된 페이지: {len(all_pages_data)}개")
    
    # 통합된 데이터를 JSON 파일로 저장
    output_file = "processed_data/all_pdfs_pages.json"
    os.makedirs("processed_data", exist_ok=True)
    
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_pages_data, f, ensure_ascii=False, indent=2)
    
    print(f"통합 데이터 저장: {output_file}")
    
    # 통계 출력
    total_chars = sum(len(page['text']) for page in all_pages_data)
    print(f"\n통계:")
    print(f"   - 총 페이지: {len(all_pages_data)}")
    print(f"   - 총 문자수: {total_chars:,}")
    print(f"   - 평균 페이지당: {total_chars // len(all_pages_data):,}자")
    
    print("\nPDF 전처리 완료!")
    print("다음 단계: python vector_store.py")

def main():
    """메인 실행 함수"""
    process_all_pdfs_in_source()

if __name__ == "__main__":
    main()